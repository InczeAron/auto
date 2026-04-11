# ============================================
# AutoScout24 Car Search – Flask Web Server
# ============================================
# Install:
#   pip install flask playwright openpyxl psycopg2-binary
#   playwright install chromium
#
# Run:
#   python app.py
# ============================================

from flask import Flask, render_template, request, jsonify, send_file, session, redirect, url_for
from playwright.sync_api import sync_playwright
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import time, os, re, threading, uuid, secrets
import psycopg2
import bcrypt

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "autoscout-fallback-key-2026")

# =========================
# DATABASE
# =========================
_conn = None

def get_db():
    global _conn
    try:
        if _conn is None or _conn.closed:
            _conn = psycopg2.connect(os.environ["DATABASE_URL"])
        else:
            # ping – ha megszakadt a kapcsolat, újracsatlakozik
            _conn.cursor().execute("SELECT 1")
    except Exception:
        _conn = psycopg2.connect(os.environ["DATABASE_URL"])
    return _conn

# Táblák létrehozása induláskor
def init_db():
    c = get_db()
    cur = c.cursor()
    cur.execute("CREATE TABLE IF NOT EXISTS used_ips (ip TEXT PRIMARY KEY)")
    c.commit()

init_db()

# =========================
# IP KEZELÉS
# =========================
def get_user_ip():
    if request.headers.get('X-Forwarded-For'):
        return request.headers.get('X-Forwarded-For').split(',')[0]
    return request.remote_addr

def has_ip(ip):
    c = get_db()
    cur = c.cursor()
    cur.execute("SELECT 1 FROM used_ips WHERE ip=%s", (ip,))
    return cur.fetchone() is not None

def save_ip(ip):
    try:
        c = get_db()
        cur = c.cursor()
        cur.execute("INSERT INTO used_ips (ip) VALUES (%s)", (ip,))
        c.commit()
    except Exception:
        get_db().rollback()

jobs = {}

BRANDS = {
    "Audi":       ["A1","A2","A3","A4","A5","A6","A7","A8","Q3","Q5","Q7","TT","R8"],
    "BMW":        ["1","2","3","4","5","6","7","X1","X3","X5","Z4","M3","M5"],
    "Mercedes-Benz":   ["A","B","C","E","S","GLA","GLC","GLE","GLK","CLA","CLS","SLK"],
    "Volkswagen": ["Golf","Polo","Passat","Tiguan","Touareg","T-Roc","ID.3","ID.4","Caddy","Sharan"],
    "Ford":       ["Focus","Fiesta","Mondeo","Kuga","Puma","Mustang","Galaxy","S-Max","Transit"],
    "Opel":       ["Astra","Corsa","Insignia","Zafira","Mokka","Crossland","Grandland"],
    "Toyota":     ["Yaris","Corolla","Camry","RAV4","C-HR","Prius","Land Cruiser","Hilux"],
    "Honda":      ["Civic","Jazz","CR-V","HR-V","Accord","FR-V"],
    "Peugeot":    ["107","206","207","208","307","308","407","508","2008","3008","5008"],
    "Renault":    ["Clio","Megane","Laguna","Kangoo","Scenic","Captur","Zoe","Kadjar"],
    "Seat":       ["Ibiza","Leon","Toledo","Altea","Arona","Ateca","Tarraco"],
    "Skoda":      ["Fabia","Octavia","Superb","Kodiaq","Karoq","Rapid","Scala"],
    "Fiat":       ["500","Punto","Bravo","Tipo","Panda","Doblo","Stilo"],
    "Kia":        ["Picanto","Rio","Ceed","Sportage","Sorento","Stonic","Niro","EV6"],
    "Hyundai":    ["i20","i30","i40","Tucson","Santa Fe","Kona","Ioniq"],
    "Mazda":      ["2","3","6","CX-3","CX-5","CX-30","MX-5"],
    "Nissan":     ["Micra","Juke","Qashqai","X-Trail","Leaf","370Z","Navara"],
    "Volvo":      ["S40","S60","S80","V40","V60","V90","XC40","XC60","XC90"],
    "Porsche":    ["911","Cayenne","Macan","Panamera","Taycan","Boxster","Cayman"],
    "Alfa Romeo": ["147","156","159","Giulia","Stelvio","MiTo","Giulietta"],
}

COUNTRIES = {
    "All Europe / Egész Európa":   "",
    "Germany / Németország":       "D",
    "Austria / Ausztria":          "A",
    "Hungary / Magyarország":      "H",
    "Italy / Olaszország":         "I",
    "France / Franciaország":      "F",
    "Spain / Spanyolország":       "E",
    "Belgium":                     "B",
    "Netherlands / Hollandia":     "NL",
    "Poland / Lengyelország":      "PL",
    "Czech Republic / Csehország": "CZ",
    "Switzerland / Svájc":         "CH",
    "Sweden / Svédország":         "S",
    "Denmark / Dánia":             "DK",
    "Portugal / Portugália":       "P",
    "Romania / Románia":           "RO",
    "Croatia / Horvátország":      "HR",
    "Luxembourg / Luxemburg":      "L",
}

# 30 perc inaktivitás után kiléptetés
ALLOWED_ORIGINS = [
    "https://aronsoft.hu",
    "https://www.aronsoft.hu",
    "http://aronsoft.hu",
    "http://www.aronsoft.hu",
]

def get_cors_origin():
    origin = request.headers.get("Origin", "")
    return origin if origin in ALLOWED_ORIGINS else ALLOWED_ORIGINS[0]

@app.route("/api/login", methods=["POST", "OPTIONS"])
def api_login():
    if request.method == "OPTIONS":
        res = app.make_default_options_response()
        res.headers["Access-Control-Allow-Origin"] = get_cors_origin()
        res.headers["Access-Control-Allow-Headers"] = "Content-Type"
        res.headers["Access-Control-Allow-Methods"] = "POST, OPTIONS"
        return res

    data = request.json or {}

    email = (data.get("email") or "").strip().lower()
    password = (data.get("password") or "").strip().encode("utf-8")

    if not email or not password:
        res = jsonify({"success": False, "error": "Hiányzó adatok"})
        res.headers["Access-Control-Allow-Origin"] = get_cors_origin()
        return res, 400

    try:
        c = get_db()
        cur = c.cursor()
        cur.execute("SELECT jelsz FROM befele WHERE felh = %s", (email,))
        row = cur.fetchone()
    except Exception as e:
        print("DB ERROR:", e)
        try:
            get_db().rollback()
        except Exception:
            pass
        res = jsonify({"success": False, "error": "DB hiba"})
        res.headers["Access-Control-Allow-Origin"] = get_cors_origin()
        return res, 500

    pw_plain = password.decode("utf-8").strip()
    print("EMAIL:", email)
    print("ROW FOUND:", row is not None)

    if row:
        stored = row[0].strip()
        print("STORED:", stored)
        try:
            match = bcrypt.checkpw(pw_plain.encode("utf-8"), stored.encode("utf-8"))
        except Exception:
            # fallback: sima szöveges összehasonlítás ha nincs bcrypt hash
            match = (pw_plain == stored)

        if match:
            session["logged_in"] = True
            res = jsonify({"success": True})
        else:
            res = jsonify({"success": False, "error": "Hibás jelszó / Wrong password"})
    else:
        res = jsonify({"success": False, "error": "Nincs ilyen user / User not found"})

    res.headers["Access-Control-Allow-Origin"] = get_cors_origin()
    return res

# Felhasználó jelszó hash generáló segédroute (csak egyszer kell, utána törölhető)
"""@app.route("/api/hash", methods=["GET"])
def api_hash():
    pw = request.args.get("pw", "")
    if not pw:
        return "?pw=yourpassword", 400
    hashed = bcrypt.hashpw(pw.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")
    return jsonify({"hash": hashed})"""

@app.route("/projects")
def projects():
    if not session.get("logged_in"):
        return redirect("/")
    return render_template("projects.html")

@app.before_request
def session_timeout():
    # Statikus és login route-ok kihagyása
    if request.endpoint in ("static",):
        return
    if "last_activity" in session:
        now = time.time()
        if now - session["last_activity"] > 1800:
            session.clear()
            # JSON kérés esetén 401, egyébként üzenet
            if request.is_json or request.path.startswith(("/search", "/status", "/download", "/models")):
                return jsonify({"error": "session_expired"}), 401
            return "❌ Session expired / Munkamenet lejárt. <a href='/'>Refresh</a>", 401
    session["last_activity"] = time.time()

# ip figyelés csak 1x lehessen belépni ip alapján
@app.route("/")
def index():
    ip = get_user_ip()
    if has_ip(ip):
        # Ha már volt bent de a session lejárt, töröljük az IP-t és engedjük újra
        if "last_activity" not in session:
            try:
                c = get_db()
                cur = c.cursor()
                cur.execute("DELETE FROM used_ips WHERE ip=%s", (ip,))
                c.commit()
            except Exception:
                c.rollback()
        else:
            return "❌ Egyszer már beléptél / You have already entered once."
    save_ip(ip)
    return render_template("index.html", brands=BRANDS, countries=list(COUNTRIES.keys()))

@app.route("/models/<brand>")
def get_models(brand):
    return jsonify(BRANDS.get(brand, []))

@app.route("/search", methods=["POST"])
def search():
    data = request.json
    job_id = str(uuid.uuid4())
    jobs[job_id] = {"status": "running", "progress": 0, "log": [], "cars": []}
    thread = threading.Thread(target=run_scrape, args=(job_id, data))
    thread.daemon = True
    thread.start()
    return jsonify({"job_id": job_id})

@app.route("/status/<job_id>")
def status(job_id):
    return jsonify(jobs.get(job_id, {}))

@app.route("/download/<job_id>")
def download(job_id):
    job = jobs.get(job_id)
    if not job or not job.get("cars"):
        return "No data / Nincs adat", 404
    filepath = os.path.join("outputs", f"{job_id}.xlsx")
    os.makedirs("outputs", exist_ok=True)
    save_to_excel(job["cars"], filepath, job["brand"], job["model"])
    return send_file(filepath, as_attachment=True,
                     download_name=f"{job['brand']}_{job['model']}_listing.xlsx")

def log(job_id, msg):
    jobs[job_id]["log"].append(msg)
    print(msg)

def extract_price(text):
    if not text:
        return None

    text = text.replace("\xa0", " ").strip()

    # 🔥 CSAK AZ ELSŐ VALID ÁR FORMÁTUM
    match = re.search(r"\d{1,3}(?:[.,\s]\d{3})+", text)

    if not match:
        return None

    number = match.group(0)

    # minden nem szám törlése
    number = re.sub(r"[^\d]", "", number)

    if not number:
        return None

    value = int(number)

    if 500 < value < 500000:
        return value

    return None

def run_scrape(job_id, data):
    brand       = data.get("brand", "")
    model       = data.get("model", "")
    year_from   = data.get("year_from") or None
    year_to     = data.get("year_to") or None
    price_from  = data.get("price_from") or None
    price_to    = data.get("price_to") or None
    country     = COUNTRIES.get(data.get("country", ""), "")
    km_from     = data.get("km_from") or None
    km_to       = data.get("km_to") or None
    seller_type = data.get("seller_type") or None

    jobs[job_id]["brand"] = brand
    jobs[job_id]["model"] = model

    brand_slug = brand.lower().replace(" ", "-")
    model_slug = model.lower().replace(" ", "-") 

    # BMW sorozat slug mapping
    BMW_SLUGS = {
        "1": "1-series-(all)",
        "2": "2-series-(all)",
        "3": "3-series-(all)",
        "4": "4-series-(all)",
        "5": "5-series-(all)",
        "6": "6-series-(all)",
        "7": "7-series-(all)",
        "8": "8-series-(all)",
        "x1": "x1", "x2": "x2", "x3": "x3",
        "x4": "x4", "x5": "x5", "x6": "x6", "x7": "x7",
        "z4": "z4", "m3": "m3", "m5": "m5",
    }
    if brand_slug == "bmw" and model.lower() in BMW_SLUGS:
        model_slug = BMW_SLUGS[model.lower()]   

    # Mercedes-Benz model slug fix
    if brand_slug == "mercedes-benz":
        MERCEDES_MODEL_MAP = {
            "a": "a-class",
            "b": "b-class",
            "c": "c-class",
            "e": "e-class",
            "s": "s-class",
            "gla": "gla-class",
            "glc": "glc-class",
            "gle": "gle-class",
            "glk": "glk-class",
            "cla": "cla-class",
            "cls": "cls-class",
            "slk": "slk-class",
        }
        model_slug = MERCEDES_MODEL_MAP.get(model_slug, model_slug)
    cars = []

    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(
                headless=True,
                args=["--disable-blink-features=AutomationControlled"]
            )
            context = browser.new_context(
                user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
                viewport={"width": 1280, "height": 800},
                locale="hu-HU",
            )
            context.add_init_script("Object.defineProperty(navigator, 'webdriver', { get: () => undefined });")
            page = context.new_page()

            for page_num in range(1, 11):
                params = f"page={page_num}"
                #if model:      params += f"&model={model.upper()}"
                if year_from:  params += f"&fregfrom={year_from}"
                if year_to:    params += f"&fregto={year_to}"
                if price_from: params += f"&pricefrom={price_from}"
                if price_to:   params += f"&priceto={price_to}"
                if country:    params += f"&cy={country}"
                if km_from:     params += f"&kmfrom={km_from}"
                if km_to:       params += f"&kmto={km_to}"
                if seller_type == "dealer":
                    params += "&atype=C&adtype=D"
                elif seller_type == "private":
                    params += "&atype=C&adtype=P"
                if year_from or year_to:
                    params += "&sort=age&desc=1"
                elif price_from or price_to:
                    params += "&sort=price&desc=0"

                url = f"https://www.autoscout24.com/lst/{brand_slug}/{model_slug}?{params}"
                log(job_id, f"📄 Loading page / Oldal betöltése: {page_num}")
                page.goto(url, wait_until="domcontentloaded", timeout=30000)

                try:
                    page.wait_for_selector("article", timeout=8000)
                except:
                    page.wait_for_selector("[data-testid='listing']", timeout=8000)

                # 🔥 görgetés, hogy betöltse az összes hirdetést
                page.mouse.wheel(0, 3000)
                time.sleep(1)
                page.mouse.wheel(0, 3000)
                time.sleep(1)

                articles = page.locator("article").all()

                if not articles:
                    articles = page.locator("[data-testid='listing']").all()

                log(job_id, f"Találatok: {len(articles)}")

                if page_num == 1:
                    for selector in ["button[id='didomi-notice-agree-button']",
                                     "button:has-text('Accept All')", "button:has-text('Accept all')"]:
                        try:
                            btn = page.locator(selector).first
                            if btn.is_visible(timeout=2000):
                                btn.click()
                                log(job_id, "✅ Cookie accepted / Cookie elfogadva")
                                time.sleep(1)
                                break
                        except Exception:
                            continue

                try:
                    page.wait_for_selector("a[href*='/offers/']", timeout=10000)
                except Exception:
                    time.sleep(3)

                

                # 🔥 fallback (új UI miatt)
                if not articles:
                    articles = page.locator("[data-testid='listing']").all()

                    log(job_id, f"  → {len(articles)} listings / hirdetés")

                    print("HTML length:", len(page.content()))

                if not articles:
                    log(job_id, "⛔ No more results / Nincs több találat")
                    break

                log(job_id, f"  → {len(articles)} listings / hirdetés")

                for article in articles:
                    try:
                        title = ""
                        try:
                            title = article.locator("h2").first.inner_text(timeout=1000).strip()
                        except Exception:
                            pass

                        price_num = None
                        price_text = ""

                        try:
                            price_text = article.locator("[class*='Price'], [class*='price']").first.inner_text(timeout=1000).strip()

                            # 🔥 LEVÁGJUK A VÉGÉRŐL A NEM SZÁM KARAKTEREKET (pl. ¹)
                            price_text = re.sub(r"[^\d€.,\s]", "", price_text)

                            price_num = extract_price(price_text)

                        except Exception:
                            pass

                        details = []
                        try:
                            spans = article.locator("dl span, [class*='VehicleDetails'] span, [class*='vehicle-detail'] span").all()
                            for s in spans[:6]:
                                t = s.inner_text(timeout=500).strip()
                                if t and t not in details:
                                    details.append(t)
                        except Exception:
                            pass

                        location = ""
                        for loc_sel in ["[data-testid='seller-address']","[data-testid='location']",
                                        "[class*='seller-address']","[class*='sellerAddress']","address",
                                        "[class*='Location']","[class*='location']","[class*='seller']"]:
                            try:
                                el = article.locator(loc_sel).first
                                if el.is_visible(timeout=500):
                                    txt = el.inner_text(timeout=500).strip()
                                    if txt and len(txt) > 2:
                                        location = txt
                                        break
                            except Exception:
                                continue
                        if not location:
                            try:
                                full_text = article.inner_text(timeout=1000)
                                match = re.search(r'\b([A-Z]{1,3}-\s*\d{4,5}[\s\S]{0,30})', full_text)
                                if match:
                                    location = match.group(1).split('\n')[0].strip()
                            except Exception:
                                pass

                        link = ""

                        try:
                            # 🔥 KÉPRE KATTINTÁS (legstabilabb módszer)
                            img_link = article.locator("a:has(img)").first

                            if img_link.count() > 0:
                                href = img_link.get_attribute("href")

                                if href:
                                    if href.startswith("/"):
                                        link = "https://www.autoscout24.com" + href
                                    else:
                                        link = href

                        except Exception:
                            pass

                        # 🔥 fallback (ha kép nem működik)
                        if not link:
                            try:
                                href = article.locator("a[href*='/offers/']").first.get_attribute("href")
                                if href:
                                    if href.startswith("/"):
                                        link = "https://www.autoscout24.com" + href
                                    else:
                                        link = href
                            except:
                                pass

                        # 🔥 tracking levágása
                        if link:
                            link = link.split("?")[0]

                        # Eladó típusának kiolvasása az article HTML-ből
                        seller_label = ""
                        try:
                            full_text = article.inner_text(timeout=1000)
                            full_html = article.inner_html(timeout=1000)
                            
                            # Debug: első article szövegének kiírása
                            #if len(cars) == 0:
                                #print("=== ARTICLE TEXT SAMPLE ===")
                                #print(full_text[:500])
                                #print("=== ARTICLE HTML SAMPLE ===")
                                #print(full_html[:800])
                                #print("===========================")                                             
                            # Seller típus keresése szövegben és HTML-ben
                            combined = full_text + full_html
                            combined_lower = combined.lower()
                            if any(x in combined_lower for x in ["private seller", "privateseller", "private-seller",
                                                                  "privatanbieter", "private_seller",
                                                                  "seller-private", "adtypeprivate"]):
                                seller_label = "private"
                            elif any(x in combined_lower for x in ["dealer", "händler",
                                                                    "adtypedealer", "seller-dealer"]):
                                seller_label = "dealer"
                        except Exception as e:
                            print(f"Seller detect error: {e}")
                            pass

                        # Szűrés eladó típusa alapján
                        if seller_type == "private" and seller_label == "dealer":
                            continue
                        if seller_type == "dealer" and seller_label == "private":
                            continue

                        if title:
                            # 🔥 MODEL SZŰRÉS (pl. GLA csak önálló szóként)
                            if model:
                                model_clean = model.lower().strip()
                                title_clean = title.lower().strip()

                                # 🔥 MERCEDES
                                if brand.lower() == "mercedes-benz":
                                    if not re.search(rf"\b{re.escape(model_clean)}\s?\d+", title_clean):
                                        continue

                                # 🔥 BMW (javítás) URL már szűr sorozatra
                                elif brand.lower() == "bmw":
                                    pass
                            # Ár megjelenítése: szám → formázott string                                       
                            price_display = f"{price_num:,} €".replace(",", ".") if price_num else price_text
                            cars.append({
                                "Cím":     title,
                                "Ár":      price_display,
                                "Ár_num":  price_num,
                                "Részletek": " | ".join(details),
                                "Helyszín": location,
                                "Link":    link
                            })
                    except Exception:
                        continue

                jobs[job_id]["progress"] = page_num * 10

            browser.close()

        cars.sort(key=lambda x: x["Ár_num"] if x["Ár_num"] else 999999)
        jobs[job_id]["cars"] = cars
        jobs[job_id]["status"] = "done"
        log(job_id, f"🎉 Done! / Kész! {len(cars)} listings / hirdetés collected.")

    except Exception as e:
        log(job_id, f"⚠️ Hiba, de megyünk tovább: {e}")
        # Ha már vannak összegyűjtött autók, azokat mentsük el
        if cars:
            cars.sort(key=lambda x: x["Ár_num"] if x["Ár_num"] else 999999)
            jobs[job_id]["cars"] = cars
            jobs[job_id]["status"] = "done"
            log(job_id, f"🎉 Részleges eredmény / Partial result: {len(cars)} listings / hirdetés.")
        else:
            jobs[job_id]["status"] = "error"
            log(job_id, "❌ Nincs eredmény / No results collected.")                         

def save_to_excel(cars, filepath, brand, model):
    wb = Workbook()
    ws = wb.active
    ws.title = f"{brand} {model}"

    headers = ["#", "Cím / Title", "Ár / Price", "Részletek / Details", "Helyszín / Location", "Link", "Ár értékelés / Price Rating"]
    header_fill = PatternFill("solid", start_color="1F3864")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # Átlagár számítás Ár_num alapján
    valid_prices = [c["Ár_num"] for c in cars if c["Ár_num"]]
    avg_price = sum(valid_prices) / len(valid_prices) if valid_prices else 0

    for i, car in enumerate(cars, 1):
        row = i + 1
        fill = PatternFill("solid", start_color="DCE6F1" if i % 2 == 0 else "FFFFFF")
        values = [i, car["Cím"], car["Ár"], car["Részletek"], car["Helyszín"], car["Link"]]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = fill
            cell.alignment = Alignment(vertical="center")
            if col == 6 and val:
                cell.hyperlink = val
                cell.value = "Open"
                cell.font = Font(name="Arial", size=10, color="0563C1", underline="single")
            else:
                cell.font = Font(name="Arial", size=10)

        # Ár értékelés – Ár_num alapján (szám!)
        price_num = car["Ár_num"]
        eval_cell = ws.cell(row=row, column=7)
        eval_cell.fill = fill
        eval_cell.alignment = Alignment(horizontal="center", vertical="center")
        if avg_price > 0 and price_num and price_num > 0:
            diff_pct = (avg_price - price_num) / avg_price * 100
            if diff_pct >= 15:
                eval_cell.value = f"✅ {diff_pct:.0f}% cheaper"
                eval_cell.font = Font(name="Arial", size=10, bold=True, color="1A7A4A")
            else:
                eval_cell.value = ""
                eval_cell.font = Font(name="Arial", size=10)
        else:
            eval_cell.value = ""
            eval_cell.font = Font(name="Arial", size=10)

    avg_row = len(cars) + 2
    avg_fill = PatternFill("solid", start_color="1F3864")
    lbl = ws.cell(row=avg_row, column=2, value="Átlagár / Average Price:")
    lbl.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    lbl.fill = avg_fill
    lbl.alignment = Alignment(horizontal="right", vertical="center")
    v = ws.cell(row=avg_row, column=3, value=f"{avg_price:,.0f}".replace(",", ".") + " €" if avg_price else "–")
    v.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    v.fill = avg_fill
    v.alignment = Alignment(horizontal="center", vertical="center")
    for col in [1, 4, 5, 6, 7]:
        ws.cell(row=avg_row, column=col).fill = avg_fill

    for col, width in enumerate([5, 35, 15, 45, 30, 10, 20], 1):
        ws.column_dimensions[ws.cell(1, col).column_letter].width = width
    ws.freeze_panes = "A2"
    wb.save(filepath)

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 8080))
    app.run(debug=False, host="0.0.0.0", port=port)
