
    #----------------------------- új verzió tisztán chat gpt-től --------------------

import time, re, smtplib, os, psycopg2
from playwright.sync_api import sync_playwright
from email.message import EmailMessage
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText


def get_db_connection():
    return psycopg2.connect(os.environ.get("DATABASE_URL"), sslmode="require")

print("DB URL:", os.environ.get("DATABASE_URL"))

def init_db():
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
    CREATE TABLE IF NOT EXISTS sent_cars (
        id SERIAL PRIMARY KEY,
        dealer_id TEXT,
        car_id TEXT,
        sent_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )
    """)

    conn.commit()
    cur.close()
    conn.close()

def load_seen(dealer_id):
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute(
        "SELECT car_id FROM sent_cars WHERE dealer_id = %s",
        (dealer_id,)
    )

    seen = {row[0] for row in cur.fetchall()}

    cur.close()
    conn.close()
    return seen

def save_seen(dealer_id, car_ids):
    if not car_ids:
        return

    conn = get_db_connection()
    cur = conn.cursor()

    for car_id in car_ids:
        cur.execute(
            "INSERT INTO sent_cars (dealer_id, car_id) VALUES (%s, %s)",
            (dealer_id, car_id)
        )

    conn.commit()
    cur.close()
    conn.close()


def send_email(subject, body, to_email, attachment=None, html=False):
    sender = os.environ.get("EMAIL_USER")
    password = os.environ.get("EMAIL_PASS")

    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = sender
    msg["To"] = to_email
    msg.set_content(body)

    # Excel csatolmány
    if attachment:
        with open(attachment, "rb") as f:
            msg.add_attachment(
                f.read(),
                maintype="application",
                subtype="octet-stream",
                filename=attachment
            )

    # Gmail SMTP
    with smtplib.SMTP("smtp.gmail.com", 587) as smtp:
        smtp.starttls()
        smtp.login(sender, password)
        smtp.send_message(msg)

    print(f"📧 Email elküldve: {to_email}")


def extract_price(text):
    if not text:
        return None
    text = text.replace("\xa0", " ").strip()
    match = re.search(r"\d{1,3}(?:[.,\s]\d{3})+", text)
    if not match:
        return None
    number = re.sub(r"[^\d]", "", match.group(0))
    if not number:
        return None
    value = int(number)
    return value if 500 < value < 500000 else None


def save_to_excel(cars, filename):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "AutoScout"

    headers = ["#", "Cím", "Ár", "Km", "Év", "Üzemanyag", "Helyszín", "Link", "Deal"]

    ws.append(headers)

    # 🔥 fejléc stílus
    header_fill = PatternFill(start_color="2F4F6F", end_color="2F4F6F", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col in range(1, len(headers)+1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # 🔥 adatok
    for car in cars:
        rating = car.get("Pontszám") or 0

        if rating > 0:
            deal_text = f"{rating}% olcsóbb"
            color = "008000"  # zöld
        else:
            deal_text = f"{rating}% drágább"
            color = "FF0000"  # piros

        ws.append([
            car.get("Sorszám"),
            car.get("Cím"),
            car.get("Ár"),
            car.get("km") or "-",
            car.get("year") or "-",
            car.get("fuel") or "-",
            car.get("Helyszín") or "-",
            "Open",
            deal_text
        ])

        row = ws.max_row

        # 🔗 link
        link_cell = ws.cell(row=row, column=8)
        link_cell.hyperlink = car.get("Link")
        link_cell.font = Font(color="0000FF", underline="single")

        # 🎯 deal színezés
        deal_cell = ws.cell(row=row, column=9)
        deal_cell.font = Font(color=color, bold=True)

    # 🔥 oszlopszélesség
    widths = [5, 50, 15, 10, 10, 10, 15]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64+i)].width = w

    wb.save(filename)
    print(f"Excel mentve: {filename}")

def build_email_html(cars):
    html = """
    <html>
    <body>
    <h2>🚗 Új autók</h2>
    <table border="1" cellpadding="6" cellspacing="0" style="border-collapse: collapse; font-family: Arial;">
        <tr style="background-color:#2f4f6f; color:white;">
            <th>#</th>
            <th>Cím</th>
            <th>Ár</th>
            <th>Km</th>
            <th>Év</th>
            <th>Link</th>
            <th>Deal</th>
        </tr>
    """
    
    for car in cars:
        rating = car.get("Pontszám") or 0

        # szín
        if rating > 0:
            color = "green"
            text = f"{rating}% olcsóbb"
        else:
            color = "red"
            text = f"{rating}% drágább"

        html += f"""
        <tr>
            <td>{car.get("Sorszám")}</td>
            <td>{car.get("Cím")}</td>
            <td>{car.get("Ár")}</td>
            <td>{car.get("km") or '-'}</td>
            <td>{car.get("year") or '-'}</td>
            <td><a href="{car.get("Link")}">Open</a></td>
            <td style="color:{color}; font-weight:bold;">{text}</td>
        </tr>
        """

    html += """
    </table>
    </body>
    </html>
    """

    return html 

"""def send_email(subject, body, to_email, attachment=None, html=False):
    sender = os.environ.get("EMAIL_USER")
    password = os.environ.get("EMAIL_PASS")

    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = sender
    msg["To"] = to_email

    if html:
        msg.add_alternative(body, subtype='html')
    else:
        msg.set_content(body)

    if attachment:
        with open(attachment, "rb") as f:
            msg.add_attachment(
                f.read(),
                maintype="application",
                subtype="octet-stream",
                filename=attachment
            )

    with smtplib.SMTP("smtp.gmail.com", 587) as smtp:
        smtp.starttls()
        smtp.login(sender, password)
        smtp.send_message(msg)"""

def run_scraper():
    print("🚀 SCRAPER START")

    searches = [
        {
            "dealer_id": "bmw_3_de_2024_2026",
            "brand": "bmw",
            "model": "3-series-(all)",
            "year_from": 2024,
            "year_to": 2026,
            "country": "D"
        },
        {
            "dealer_id": "honda_jazz_at_2020_2026",
            "brand": "honda",
            "model": "jazz",
            "year_from": 2020,
            "year_to": 2026,
            "country": "A"
        }
    ]

    for search in searches:
        print(f"\n🔍 Keresés: {search['dealer_id']}")

        brand = search["brand"]
        model_slug = search["model"]
        year_from = search["year_from"]
        year_to = search["year_to"]
        country = search["country"]
        dealer_id = search["dealer_id"]

        cars = []  # 🔥 MINDEN kereséshez új lista

        with sync_playwright() as p:
            browser = p.chromium.launch(
                headless=True,
                args=[
                    "--disable-blink-features=AutomationControlled",
                    "--no-sandbox",
                    "--disable-dev-shm-usage"
                ]
            )

        context = browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120 Safari/537.36",
            viewport={"width": 1280, "height": 800},
            locale="de-DE"
        )

        # 🔥 stealth hack
        context.add_init_script("""
        Object.defineProperty(navigator, 'webdriver', { get: () => undefined });
        """)

        page = context.new_page()

        for page_num in range(1, 4):  # max 3 oldal
            print(f"\n📄 OLDAL: {page_num}")

            url = f"https://www.autoscout24.com/lst/{brand}/{model_slug}?page={page_num}&fregfrom={year_from}&fregto={year_to}&cy={country}"

            try:
                page.goto(url, timeout=20000)
            except:
                print("❌ Nem tölt be az oldal")
                break

            # DEBUG
            html_len = len(page.content())
            print("HTML méret:", html_len)

            if html_len < 5000:
                print("⛔ VALÓSZÍNŰ BLOCK / CAPTCHA")
                page.screenshot(path=f"debug_block_{page_num}.png")
                break

            # cookie accept
            try:
                btn = page.locator("button:has-text('Accept')").first
                if btn.is_visible(timeout=2000):
                    btn.click()
                    print("✅ Cookie OK")
            except:
                pass

            try:
                page.wait_for_selector("article", timeout=5000)
            except:
                print("⚠️ Nincs article → fallback")

            articles = page.locator("article").all()

            if not articles:
                print("⛔ NINCS TALÁLAT → STOP")
                break

            print("Találatok:", len(articles))

            for i, article in enumerate(articles):
                if i > 30:  # limit
                    break

                year = ""
                fuel = ""
                location = ""
                rating = 0
                km = None

                try:
                    title = article.locator("h2").first.inner_text(timeout=1000)

                    price_text = article.locator("[class*='Price']").first.inner_text(timeout=1000)
                    price_num = extract_price(price_text)

                    link = article.locator("a[href*='/offers/']").first.get_attribute("href")
                    if link and link.startswith("/"):
                        link = "https://www.autoscout24.com" + link

                    cars.append({
                        "Sorszám": len(cars) + 1,
                        "Cím": title,
                        "Ár": f"{price_num:,} €".replace(",", ".") if price_num else price_text,
                        "Ár_num": price_num,
                        "Km": km,
                        "Év": year,
                        "Üzemanyag": fuel,
                        "Helyszín": location,
                        "Link": link,
                        "Pontszám": rating
                    })

                    if not price_num:
                        continue

                    details = article.locator("span").all()

                    for d in details:
                        txt = d.inner_text(timeout=500).strip().lower()

                        if "km" in txt:
                            km = re.sub(r"[^\d]", "", txt)
                            cars[-1]["km"] = int(km) if km else None

                        elif re.search(r"\d{2}/\d{4}", txt):
                            cars[-1]["year"] = txt

                        elif "diesel" in txt or "benzin" in txt or "gasoline" in txt:
                            cars[-1]["fuel"] = txt

                     # 🔥 HELYSZÍN KINYERÉS
                    try:
                        loc = article.locator("[class*='Location'], [class*='location'], [class*='seller']").first.inner_text(timeout=500)
                        cars[-1]["Helyszín - Location"] = loc.strip()
                    except:
                        cars[-1]["Helyszín"] = ""

                except:
                    continue

            time.sleep(2)  # 🔥 ne pörögjön túl gyorsan

        browser.close()

        # 🔥 Átlag ár számítás
        valid_prices = [c["Ár_num"] for c in cars if c.get("Ár_num")]
        avg_price = sum(valid_prices) / len(valid_prices) if valid_prices else 0

        # 🔥 Pontszám számítás
        for c in cars:
            if c.get("Ár_num") and avg_price:
                diff = (avg_price - c["Ár_num"]) / avg_price * 100
                c["Pontszám"] = round(diff)
            else:
                c["Pontszám"] = None

        # rendezés
        cars.sort(key=lambda x: x.get("Pontszám") or -999, reverse=True)

        print(f"\n🎯 Talált autók: {len(cars)}")

        print(f"Talált autók száma: {len(cars)}")

        seen = load_seen(dealer_id)

        new_cars = []
        new_ids = []

        for car in cars:
            link = car.get("Link")

        if not link:
            continue

        car_id = link.split("/")[-1].split("?")[0]

        if car_id not in seen:
            new_cars.append(car)
            new_ids.append(car_id)

        if not new_cars:
            print("Nincs új autó")
            send_email(
                subject="🚗 AutoScout – nincs új autó - not a new car",
                body="A mai futás során nem találtunk új hirdetéseket.",
                to_email=["aronincze@aronsoft.hu", "inczearon@gmail.com"]
            )
            continue

        # rendezés
        new_cars.sort(key=lambda x: x.get("Pontszám - Score") or -999, reverse=True)

        # Excel
        filename = f"{dealer_id}.xlsx"
        save_to_excel(new_cars, filename)

        # HTML email
        email_html = build_email_html(new_cars)

        # DB mentés
        save_seen(dealer_id, new_ids)

        send_email(
            subject=f"🚗 {len(new_cars)} új autó - new car (AutoScout)",
            body=email_html,
            to_email=["aronincze@aronsoft.hu", "inczearon@gmail.com"],
            html=True,
            attachment=filename
        )

        

    for car in cars[:5]:
        print(car)   

    
if __name__ == "__main__":
    try:
        run_scraper()
    except Exception as e:
        send_email(
            subject="❌ SCRAPER HIBA",
            body=str(e),
            to_email="aronincze@aronsoft.hu"
        )