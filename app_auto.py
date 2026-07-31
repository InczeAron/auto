import time
import re
import smtplib
import os
import json
import random
import hashlib  # ← EZ KELL AZ UJJLENYOMATHOZ
from pathlib import Path
from email.message import EmailMessage
from playwright.sync_api import sync_playwright

# =========================
# SEEN LINKS (JSON FÁJL KEZELÉS)
# =========================
SEEN_DIR = Path("seen_links")
SEEN_DIR.mkdir(exist_ok=True)

def load_seen_links(dealer_id):
    file = SEEN_DIR / f"{dealer_id}.json"
    if not file.exists():
        return set()
    with open(file, "r", encoding="utf-8") as f:
        try:
            return set(json.load(f))
        except json.JSONDecodeError:
            return set()

def save_seen_links(dealer_id, links):
    file = SEEN_DIR / f"{dealer_id}.json"
    with open(file, "w", encoding="utf-8") as f:
        json.dump(sorted(list(links)), f, indent=2)

# =========================
# EMAIL
# =========================
def send_email(subject, body, to_email, attachment=None, html=False):
    sender = os.environ.get("EMAIL_USER")
    password = os.environ.get("EMAIL_PASS")

    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = sender
    msg["To"] = ", ".join(to_email) if isinstance(to_email, list) else to_email

    if html:
        msg.add_alternative(body, subtype="html")
    else:
        msg.set_content(body)

    if attachment and os.path.exists(attachment):
        with open(attachment, "rb") as f:
            msg.add_attachment(
                f.read(), maintype="application", subtype="octet-stream",
                filename=os.path.basename(attachment)
            )

    with smtplib.SMTP_SSL("smtp.forpsi.com", 465) as smtp:
        smtp.login(sender, password)
        smtp.send_message(msg)
    print(f"  📧 Email elküldve: {subject}")

# =========================
# EXCEL - MEDIÁNNAL KIEGÉSZÍTVE
# =========================
def save_to_excel(cars, filename, medians=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    if not medians: medians = {}

    wb = Workbook()
    ws = wb.active
    ws.title = "AutoScout"

    headers = ["#", "Title", "Price", "Mileage", "Year", "Fuel", "Location", "Link", "Deal"]
    header_fill = PatternFill(start_color="2F4F6F", end_color="2F4F6F", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for car in cars:
        rating = car.get("Pontszám") or 0
        deal_text = f"{abs(rating)}% cheaper" if rating > 0 else f"{abs(rating)}% more expensive"
        color = "008000" if rating > 0 else "FF0000"

        ws.append([
            car.get("Sorszám"), car.get("Cím"), car.get("Ár"), car.get("Km") or "-",
            car.get("Év") or "-", car.get("Üzemanyag") or "-", car.get("Helyszín") or "-",
            "Open", deal_text
        ])

        row = ws.max_row
        link_cell = ws.cell(row=row, column=8)
        link_cell.hyperlink = car.get("Link") or ""
        link_cell.font = Font(color="0000FF", underline="single")

        deal_cell = ws.cell(row=row, column=9)
        deal_cell.font = Font(color=color, bold=True)

    widths = [5, 50, 15, 12, 12, 12, 20, 8, 15]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = w

    # 📊 MEDIÁN RÉSZ HOZZÁADÁSA AZ EXCELHEZ
    if medians:
        median_row = len(cars) + 3
        med_fill = PatternFill(start_color="2F4F6F", end_color="2F4F6F", fill_type="solid")
        med_font_white = Font(bold=True, color="FFFFFF")
        med_font_blue = Font(bold=True, color="1F3864")

        title_cell = ws.cell(row=median_row, column=2, value="Medián ár évjárat szerint:")
        title_cell.font = med_font_white
        title_cell.fill = med_fill
        title_cell.alignment = Alignment(horizontal="right")
        ws.cell(row=median_row, column=1).fill = med_fill

        for i, (year, median) in enumerate(sorted(medians.items(), reverse=True)):
            r = median_row + 1 + i
            
            yr_cell = ws.cell(row=r, column=2, value=year)
            yr_cell.font = med_font_blue
            yr_cell.alignment = Alignment(horizontal="right")
            
            val_cell = ws.cell(row=r, column=3, value=f"{int(median):,} €".replace(",", "."))
            val_cell.font = med_font_blue
            val_cell.alignment = Alignment(horizontal="center")

    wb.save(filename)
    print(f"  📁 Excel mentve: {filename}")

# =========================
# HTML EMAIL - MEDIÁNNAL KIEGÉSZÍTVE
# =========================
def build_email_html(cars, medians, search_label=""):
    html = f"""
    <html><body>
    <h2>🚗 {search_label}</h2>
    <table border="1" cellpadding="6" cellspacing="0" style="border-collapse:collapse;font-family:Arial;">
        <tr style="background-color:#2f4f6f;color:white;">
            <th>#</th><th>Title</th><th>Price</th><th>Mileage</th><th>Year</th><th>Location</th><th>Link</th><th>Deal</th>
        </tr>
    """
    for car in cars:
        rating = car.get("Pontszám") or 0
        color = "green" if rating > 0 else "red"
        text = f"{abs(rating)}% cheaper" if rating > 0 else f"{abs(rating)}% more expensive"
        html += f"""
        <tr>
            <td>{car.get("Sorszám")}</td>
            <td>{car.get("Cím")}</td>
            <td>{car.get("Ár")}</td>
            <td>{car.get("Km") or '-'}</td>
            <td>{car.get("Év") or '-'}</td>
            <td>{car.get("Helyszín") or '-'}</td>
            <td><a href="{car.get("Link")}">Open</a></td>
            <td style="color:{color};font-weight:bold;">{text}</td>
        </tr>"""
    html += "</table>"
    
    # 📊 MEDIÁN RÉSZ HOZZÁADÁSA AZ EMAILHEZ
    if medians:
        html += """<br><table border="0" cellpadding="5" style="font-family:Arial; border: 2px solid #2f4f6f; border-radius: 5px; width: 300px;">
                    <tr style="background-color:#2f4f6f; color:white;">
                        <th colspan="2" style="text-align:left;">📊 Medián árak évjárat szerint</th>
                    </tr>"""
        for year, median in sorted(medians.items(), reverse=True):
            formatted_median = f"{int(median):,} €".replace(",", ".")
            html += f"""<tr style="border-bottom: 1px solid #eee;">
                            <td style="font-weight:bold; color:#1F3864; text-align:right; padding-right:10px;">{year}</td>
                            <td style="color:#333;">{formatted_median}</td>
                        </tr>"""
        html += "</table>"

    html += "</body></html>"
    return html

# =========================
# PRICE EXTRACT
# =========================
def extract_price(text):
    if not text: return None
    text = text.replace("\xa0", " ").strip()
    match = re.search(r"\d{1,3}(?:[.,\s]\d{3})+", text)
    if not match: match = re.search(r"\d+", text)
    if not match: return None
    value = int(re.sub(r"[^\d]", "", match.group(0)))
    return value if 500 < value < 500000 else None

# =========================
# LINK EXTRACT - PÁNCÉLOZOTT SHADOW DOM + CTRL+KATTINTÁS
# =========================
def get_real_link(context, page, article):
    try:
        link_elem = article.locator("a").first
        href = link_elem.get_attribute("href", timeout=1000)
        if href and "/lst/" not in href and not href.startswith("javascript"):
            if href.startswith("/"):
                return f"https://www.autoscout24.com{href}"
            elif href.startswith("http"):
                return href
    except:
        pass

    try:
        with context.expect_page(timeout=5000) as new_page_info:
            article.click(modifiers=["ControlOrMeta"])
            
        new_page = new_page_info.value
        time.sleep(1)
        real_url = new_page.url
        new_page.close()
        
        if "autoscout24.com" in real_url and len(real_url) > 40:
            return real_url
    except:
        try:
            page.go_back(wait_until="domcontentloaded", timeout=5000)
        except:
            pass

    return ""

# =========================
# AUTO UJJLENYOMAT KÉSZÍTÉSE (Duplikáció szűrésére)
# =========================
def get_car_fingerprint(link, title, price_num, km):
    """
    Kinyeri az autó egyedi ID-ját a linkből. 
    Ha nem talál ID-t, akkor a Cím+Km+Ár alapján generál egy egyedi kódot.
    """
    # 1. Próbáljuk megkeresni az AutoScout GUID azonosítót
    guid_match = re.search(r'([a-f0-9]{8}-[a-f0-9]{4}-[a-f0-9]{4}-[a-f0-9]{4}-[a-f0-9]{12})', link, re.I)
    if guid_match:
        return f"ID_{guid_match.group(1)}"

    # 2. Próbáljuk megkeresni a numerikus azonosítót a link végén
    num_match = re.search(r'(\d{7,})', link)
    if num_match:
        return f"ID_{num_match.group(1)}"

    # 3. HA NINCS KINYERHETŐ ID: Cím + Kilométer + Ár alapján csinálunk egy "ujjlenyomatot"
    raw_string = f"{title}_{km}_{price_num}".strip().lower()
    hash_code = hashlib.md5(raw_string.encode('utf-8')).hexdigest()
    return f"FP_{hash_code}"

# =========================
# SCRAPE ONE SEARCH - PÁNCÉLOZOTT ADATKINYERÉS
# =========================
def scrape_search(page, context, brand, model_slug, year_from, year_to, country):
    cars = []
    
    for page_num in range(1, 10):
        url = (f"https://www.autoscout24.com/lst/{brand}/{model_slug}"
               f"?page={page_num}&fregfrom={year_from}&fregto={year_to}&cy={country}")
        print(f"  📄 Oldal: {page_num}")
        
        try:
            page.goto(url, timeout=25000)
        except Exception as e:
            print("❌ page.goto() hiba:", e)
            break

        if len(page.content()) < 5000:
            print("  ⛔ Valószínű CAPTCHA/block")
            break

        try:
            btn = page.locator("button:has-text('Accept')").first
            if btn.is_visible(timeout=2000): btn.click()
        except: pass

        try: page.wait_for_selector("article", timeout=5000)
        except: pass

        articles = page.locator("article").all()
        if not articles:
            print("⛔ Nincs találat!")
            break

        print(f"  → {len(articles)} hirdetés vizsgálata...")

        for article in articles:
            try:
                link = get_real_link(context, page, article)
                if not link:
                    continue

                title = ""
                try:
                    title = article.locator("h2").first.inner_text(timeout=500).strip()
                except:
                    try:
                        title = article.locator("h3").first.inner_text(timeout=500).strip()
                    except:
                        try:
                            title = article.locator("a").first.inner_text(timeout=500).strip()
                        except:
                            pass 

                price_text = ""
                price_num = None
                try:
                    price_text = article.locator("[class*='Price']").first.inner_text(timeout=500).strip()
                    price_num = extract_price(price_text)
                except:
                    try:
                        all_text = article.inner_text(timeout=500)
                        match = re.search(r'€\s*[\d.,]+', all_text)
                        if match:
                            price_text = match.group(0)
                            price_num = extract_price(price_text)
                    except:
                        pass

                km, year, fuel, location = None, "", "", ""
                try:
                    for d in article.locator("span").all():
                        txt = d.inner_text(timeout=300).strip().lower()
                        if "km" in txt and re.search(r"\d", txt):
                            km_str = re.sub(r"[^\d]", "", txt)
                            km = int(km_str) if km_str else None
                        elif re.search(r"\d{2}/\d{4}", txt): year = txt
                        elif any(f in txt for f in ["diesel", "benzin", "gasoline", "petrol", "electric", "hybrid"]): fuel = txt
                except: pass

                try:
                    for s in article.locator("span").all():
                        txt = s.inner_text(timeout=200).strip()
                        if re.search(r"[A-Z]{2}-\d{4,5}", txt):
                            location = txt
                            break
                except: pass

                cars.append({
                    "Sorszám": len(cars) + 1, "Cím": title,
                    "Ár": f"{price_num:,} €".replace(",", ".") if price_num else price_text,
                    "Ár_num": price_num, "Km": km, "Év": year, "Üzemanyag": fuel,
                    "Helyszín": location, "Link": link, "Pontszám": 0
                })
            except:
                continue
        
        time.sleep(2)
    
    print(f"  ✅ Összesen {len(cars)} db autó mentve.")
    return cars

# =========================
# MAIN
# =========================
def run_scraper():
    print("🚀 SCRAPER START")

    dealers = [
        {
            "dealer_id": "dealer1",
            "emails": ["aronincze@aronsoft.hu"],
            "searches": [
                {"brand": "bmw",  "model": "3-series-(all)", "year_from": 2024, "year_to": 2026, "country": "D"},
                {"brand": "audi", "model": "a6",             "year_from": 2024, "year_to": 2026, "country": "A"},
            ]
        },
        {
            "dealer_id": "dealer2",
            "emails": ["inczearon@gmail.com"],
            "searches": [
                {"brand": "mercedes-benz", "model": "gla-(all)", "year_from": 2024, "year_to": 2026, "country": "D"},
                {"brand": "volkswagen",    "model": "golf",      "year_from": 2024, "year_to": 2026, "country": "D"},
            ]
        },
    ]

    USER_AGENTS = [
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120 Safari/537.36",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/119 Safari/537.36",
    ]

    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=True,
            args=["--disable-blink-features=AutomationControlled", "--no-sandbox", "--disable-dev-shm-usage"]
        )

        for dealer in dealers:
            dealer_id = dealer["dealer_id"]
            emails    = dealer["emails"]

            print(f"\n{'='*40}\n🏢 Dealer: {dealer_id}")

            context = browser.new_context(
                user_agent=random.choice(USER_AGENTS),
                viewport={"width": 1280, "height": 800},
                locale="de-DE"
            )
            context.add_init_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined});")

            seen_links = load_seen_links(dealer_id)
            print(f"🔍 Ismert autók száma: {len(seen_links)}")

            for search in dealer["searches"]:
                label = f"{search['brand']} {search['model']} ({search['country']})"
                print(f"\n🔍 Keresés: {label}")

                page = context.new_page()
                cars = scrape_search(page, context, search["brand"], search["model"], search["year_from"], search["year_to"], search["country"])
                page.close()

                # Medián számítás
                cars_by_year, medians = {}, {}
                for c in cars:
                    year, price = c.get("Év"), c.get("Ár_num")
                    if year and price:
                        y = year.split("/")[-1]
                        cars_by_year.setdefault(y, []).append(price)

                for y, prices in cars_by_year.items():
                    prices.sort()
                    n = len(prices)
                    medians[y] = prices[n//2] if n % 2 == 1 else (prices[n//2 - 1] + prices[n//2]) / 2

                if medians:
                    print("  📊 Évenkénti medián árak:")
                    for y, m in sorted(medians.items(), reverse=True):
                        print(f"     - {y}: {int(m):,} €".replace(",", "."))

                # Pontszám számítás
                for c in cars:
                    year, price = c.get("Év"), c.get("Ár_num")
                    if year and price:
                        y = year.split("/")[-1]
                        median = medians.get(y)
                        if median:
                            c["Pontszám"] = round((median - price) / median * 100)

                # KIVÁLOGATÁS: Csak az ÚJ autók (Ujjlenyomat alapján!)
                new_cars = []
                for car in cars:
                    link = car.get("Link")
                    if not link: continue
                    
                    # Ujjlenyomat generálása
                    fingerprint = get_car_fingerprint(
                        link, 
                        car.get("Cím", ""), 
                        car.get("Ár_num"), 
                        car.get("Km")
                    )
                    
                    # Csak akkor tesszük a listába, ha ezt az ujjlenyomatot még nem láttuk
                    if fingerprint not in seen_links:
                        seen_links.add(fingerprint)
                        new_cars.append(car)

                print(f"📬 Ebből {len(new_cars)} db ÚJ autó")

                if not new_cars:
                    continue

                new_cars.sort(key=lambda x: (parse_date(x.get("Év")), x.get("Pontszám") or -999), reverse=True)
                for i, c in enumerate(new_cars, 1):
                    c["Sorszám"] = i

                safe_model = search['model'].replace(" ", "_")
                filename = f"{dealer_id}_{search['brand']}_{safe_model}.xlsx"
                
                save_to_excel(new_cars, filename, medians)

                email_html = build_email_html(new_cars, medians, label)
                send_email(
                    subject=f"🚗 {len(new_cars)} új {search['brand']} {search['model']} – {dealer_id}",
                    body=email_html,
                    to_email=emails,
                    html=True,
                    attachment=filename
                )

            save_seen_links(dealer_id, seen_links)
            print("💾 Seen linkek frissítve a JSON fájlban.")
            context.close()

        browser.close()
    print("\n✅ SCRAPER KÉSZ")

def parse_date(date_str):
    try:
        month, year = date_str.split("/")
        return int(year), int(month)
    except:
        return (0, 0)

if __name__ == "__main__":
    try:
        run_scraper()
    except Exception as e:
        import traceback
        send_email(
            subject="❌ SCRAPER HIBA",
            body=traceback.format_exc(),
            to_email="aronincze@aronsoft.hu"
        )