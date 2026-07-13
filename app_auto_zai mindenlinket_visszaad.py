import time
import re
import smtplib
import os
import json
import random
from pathlib import Path
from email.message import EmailMessage
from playwright.sync_api import sync_playwright

# =========================
# SEEN LINKS (JSON FÁJL KEZELÉS) - Z.AI
# =========================
SEEN_DIR = Path("seen_links")
SEEN_DIR.mkdir(exist_ok=True)

def load_seen_links(dealer_id):
    """Betölti az már látott linkeket a JSON fájlból."""
    file = SEEN_DIR / f"{dealer_id}.json"
    if not file.exists():
        return set()
    with open(file, "r", encoding="utf-8") as f:
        try:
            return set(json.load(f))
        except json.JSONDecodeError:
            return set()

def save_seen_links(dealer_id, links):
    """Menti a teljes link halmazt vissza a JSON fájlba."""
    file = SEEN_DIR / f"{dealer_id}.json"
    with open(file, "w", encoding="utf-8") as f:
        json.dump(sorted(list(links)), f, indent=2)

# =========================
# EMAIL
# =========================
def send_email(subject, body, to_email, attachment=None, html=False):
    sender = os.environ.get("EMAIL_USER")
    password = os.environ.get("EMAIL_PASS")

    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = sender
    msg["To"] = ", ".join(to_email) if isinstance(to_email, list) else to_email

    if html:
        msg.add_alternative(body, subtype="html")
    else:
        msg.set_content(body)

    if attachment and os.path.exists(attachment):
        with open(attachment, "rb") as f:
            msg.add_attachment(
                f.read(), maintype="application", subtype="octet-stream",
                filename=os.path.basename(attachment)
            )

    with smtplib.SMTP_SSL("smtp.forpsi.com", 465) as smtp:
        smtp.login(sender, password)
        smtp.send_message(msg)
    print(f"📧 Email elküldve: {msg['To']}")

# =========================
# EXCEL
# =========================
def save_to_excel(cars, filename):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "AutoScout"

    headers = ["#", "Title", "Price", "Mileage", "Year", "Fuel", "Location", "Link", "Deal"]
    header_fill = PatternFill(start_color="2F4F6F", end_color="2F4F6F", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for car in cars:
        rating = car.get("Pontszám") or 0
        deal_text = f"{abs(rating)}% cheaper" if rating > 0 else f"{abs(rating)}% more expensive"
        color = "008000" if rating > 0 else "FF0000"

        ws.append([
            car.get("Sorszám"), car.get("Cím"), car.get("Ár"), car.get("Km") or "-",
            car.get("Év") or "-", car.get("Üzemanyag") or "-", car.get("Helyszín") or "-",
            "Open", deal_text
        ])

        row = ws.max_row
        link_cell = ws.cell(row=row, column=8)
        link_cell.hyperlink = car.get("Link") or ""
        link_cell.font = Font(color="0000FF", underline="single")

        deal_cell = ws.cell(row=row, column=9)
        deal_cell.font = Font(color=color, bold=True)

    widths = [5, 50, 15, 12, 12, 12, 20, 8, 15]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = w

    wb.save(filename)
    print(f"✅ Excel mentve: {filename}")

# =========================
# HTML EMAIL
# =========================
def build_email_html(cars, medians, search_label=""):
    html = f"""
    <html><body>
    <h2>🚗 New cars – {search_label}</h2>
    <table border="1" cellpadding="6" cellspacing="0" style="border-collapse:collapse;font-family:Arial;">
        <tr style="background-color:#2f4f6f;color:white;">
            <th>#</th><th>Title</th><th>Price</th><th>Mileage</th><th>Year</th><th>Location</th><th>Link</th><th>Deal</th>
        </tr>
    """
    for car in cars:
        rating = car.get("Pontszám") or 0
        color = "green" if rating > 0 else "red"
        text = f"{abs(rating)}% cheaper" if rating > 0 else f"{abs(rating)}% more expensive"
        html += f"""
        <tr>
            <td>{car.get("Sorszám")}</td>
            <td>{car.get("Cím")}</td>
            <td>{car.get("Ár")}</td>
            <td>{car.get("Km") or '-'}</td>
            <td>{car.get("Év") or '-'}</td>
            <td>{car.get("Helyszín") or '-'}</td>
            <td><a href="{car.get("Link")}">Open</a></td>
            <td style="color:{color};font-weight:bold;">{text}</td>
        </tr>"""
    html += "</table>"
    
    if medians:
        html += "<br><h3>📊 Median prices by year</h3><ul>"
        for year, median in sorted(medians.items(), reverse=True):
            html += f"<li>{year}: {int(median):,} €</li>".replace(",", ".")
        html += "</ul>"

    html += "</body></html>"
    return html

# =========================
# PRICE EXTRACT
# =========================
def extract_price(text):
    if not text: return None
    text = text.replace("\xa0", " ").strip()
    match = re.search(r"\d{1,3}(?:[.,\s]\d{3})+", text)
    if not match: match = re.search(r"\d+", text)
    if not match: return None
    value = int(re.sub(r"[^\d]", "", match.group(0)))
    return value if 500 < value < 500000 else None

# =========================
# LINK EXTRACT - JAVÍTOTT VERZIÓ
# =========================
def extract_link(article):
    """Kinyeri a linket az article elemből - több módszerrel próbálkozik."""
    link = ""
    
    # 1. MÓDSZER: <a> tag href attribútuma
    try:
        anchors = article.locator("a[href*='/offers/']").all()
        if anchors:
            href = anchors[0].get_attribute("href", timeout=1000)
            if href:
                if href.startswith("/"):
                    link = f"https://www.autoscout24.com{href}"
                elif href.startswith("http"):
                    link = href
    except:
        pass
    
    # 2. MÓDSZER: Bármilyen <a> tag az article-ban
    if not link:
        try:
            anchor = article.locator("a").first
            if anchor:
                href = anchor.get_attribute("href", timeout=1000)
                if href:
                    if "/offers/" in href:
                        if href.startswith("/"):
                            link = f"https://www.autoscout24.com{href}"
                        elif href.startswith("http"):
                            link = href
        except:
            pass
    
    # 3. MÓDSZER: data-ad-id attribútum (AutoScout új formátum)
    if not link:
        try:
            ad_id = article.get_attribute("data-ad-id", timeout=1000)
            if ad_id:
                link = f"https://www.autoscout24.com/offers/{ad_id}"
        except:
            pass
    
    # 4. MÓDSZER: Régi guid regex az outerHTML-ben
    if not link:
        try:
            article_html = article.evaluate("e => e.outerHTML")
            guid_match = re.search(r'guid=([a-f0-9-]{36})', article_html)
            if guid_match:
                link = f"https://www.autoscout24.com/offers/{guid_match.group(1)}"
        except:
            pass
    
    # 5. MÓDSZER: /offers/ URL regex az outerHTML-ben
    if not link:
        try:
            article_html = article.evaluate("e => e.outerHTML")
            offers_match = re.search(r'href="(/offers/[^"]+)"', article_html)
            if offers_match:
                link = f"https://www.autoscout24.com{offers_match.group(1)}"
        except:
            pass
    
    # 6. MÓDSZER: ID alapú URL az outerHTML-ben (pl. data-id, id="ad-123")
    if not link:
        try:
            article_html = article.evaluate("e => e.outerHTML")
            # Keresünk olyan számokat, amik lehetnek ad ID-k
            id_match = re.search(r'data-[a-z-]*id["\s]*=["\s]*["\']?(\d{6,})', article_html, re.I)
            if id_match:
                link = f"https://www.autoscout24.com/offers/{id_match.group(1)}"
        except:
            pass
    
    return link

# =========================
# SCRAPE ONE SEARCH
# =========================
def scrape_search(page, brand, model_slug, year_from, year_to, country):
    cars = []
    no_link_count = 0
    
    for page_num in range(1, 10):
        print(f"  📄 Oldal: {page_num}")
        url = (f"https://www.autoscout24.com/lst/{brand}/{model_slug}"
               f"?page={page_num}&fregfrom={year_from}&fregto={year_to}&cy={country}")
        try:
            page.goto(url, timeout=25000)
        except Exception as e:
            print("❌ page.goto() hiba:", e)
            break

        if len(page.content()) < 5000:
            print("  ⛔ Valószínű CAPTCHA/block")
            break

        try:
            btn = page.locator("button:has-text('Accept')").first
            if btn.is_visible(timeout=2000): btn.click()
        except: pass

        try: page.wait_for_selector("article", timeout=5000)
        except: pass

        articles = page.locator("article").all()
        if not articles:
            print("⛔ Nincs találat!")
            break

        print(f"  → {len(articles)} hirdetés")

        for article in articles:
            try:
                title = article.locator("h2").first.inner_text(timeout=1000).strip()
                price_text = article.locator("[class*='Price']").first.inner_text(timeout=1000).strip()
                price_num = extract_price(price_text)

                # JAVÍTOTT LINK KINYERÉS
                link = extract_link(article)
                
                if not link:
                    no_link_count += 1
                    # DEBUG: Kiírjuk az első pár link nélküli autó HTML-jét
                    if no_link_count <= 2:
                        try:
                            sample_html = article.evaluate("e => e.outerHTML")[:500]
                            print(f"  ⚠️ Nincs link - HTML minta: {sample_html}...")
                        except:
                            pass
                
                km, year, fuel, location = None, "", "", ""
                try:
                    for d in article.locator("span").all():
                        txt = d.inner_text(timeout=300).strip().lower()
                        if "km" in txt and re.search(r"\d", txt):
                            km_str = re.sub(r"[^\d]", "", txt)
                            km = int(km_str) if km_str else None
                        elif re.search(r"\d{2}/\d{4}", txt): year = txt
                        elif any(f in txt for f in ["diesel", "benzin", "gasoline", "petrol", "electric", "hybrid"]): fuel = txt
                except: pass

                try:
                    for s in article.locator("span").all():
                        txt = s.inner_text(timeout=200).strip()
                        if re.search(r"[A-Z]{2}-\d{4,5}", txt):
                            location = txt
                            break
                except: pass

                cars.append({
                    "Sorszám": len(cars) + 1, "Cím": title,
                    "Ár": f"{price_num:,} €".replace(",", ".") if price_num else price_text,
                    "Ár_num": price_num, "Km": km, "Év": year, "Üzemanyag": fuel,
                    "Helyszín": location, "Link": link, "Pontszám": 0
                })
            except: continue
        time.sleep(2)
    
    if no_link_count > 0:
        print(f"  ⚠️ Összesen {no_link_count} autónak nem sikerült linket kinyerni!")
    
    return cars

# =========================
# MAIN
# =========================
def run_scraper():
    print("🚀 SCRAPER START")

    dealers = [
        {
            "dealer_id": "dealer1",
            "emails": ["aronincze@aronsoft.hu"],
            "searches": [
                {"brand": "bmw",  "model": "3-series-(all)", "year_from": 2024, "year_to": 2026, "country": "D"},
                {"brand": "audi", "model": "a6",             "year_from": 2024, "year_to": 2026, "country": "A"},
            ]
        },
        {
            "dealer_id": "dealer2",
            "emails": ["inczearon@gmail.com"],
            "searches": [
                {"brand": "mercedes-benz", "model": "gla-(all)", "year_from": 2024, "year_to": 2026, "country": "D"},
                {"brand": "volkswagen",    "model": "golf",      "year_from": 2024, "year_to": 2026, "country": "D"},
            ]
        },
    ]

    USER_AGENTS = [
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120 Safari/537.36",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/119 Safari/537.36",
    ]

    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=True,
            args=["--disable-blink-features=AutomationControlled", "--no-sandbox", "--disable-dev-shm-usage"]
        )

        for dealer in dealers:
            dealer_id = dealer["dealer_id"]
            emails    = dealer["emails"]

            print(f"\n{'='*40}\n🏢 Dealer: {dealer_id}")

            context = browser.new_context(
                user_agent=random.choice(USER_AGENTS),
                viewport={"width": 1280, "height": 800},
                locale="de-DE"
            )
            context.add_init_script("Object.defineProperty(navigator, 'webdriver', { get: () => undefined });")

            seen_links = load_seen_links(dealer_id)
            print(f"🔍 Betöltött látott linkek száma: {len(seen_links)}")

            all_new_cars = []
            all_new_links = []
            all_medians = {}

            for search in dealer["searches"]:
                label = f"{search['brand']} {search['model']} ({search['country']})"
                print(f"\n🔍 Keresés: {label}")

                page = context.new_page()
                cars = scrape_search(page, search["brand"], search["model"], search["year_from"], search["year_to"], search["country"])
                page.close()

                # Medián számítás
                cars_by_year, medians = {}, {}
                for c in cars:
                    year, price = c.get("Év"), c.get("Ár_num")
                    if year and price:
                        y = year.split("/")[-1]
                        cars_by_year.setdefault(y, []).append(price)

                for y, prices in cars_by_year.items():
                    prices.sort()
                    n = len(prices)
                    medians[y] = prices[n//2] if n % 2 == 1 else (prices[n//2 - 1] + prices[n//2]) / 2
                
                all_medians.update(medians)

                # Pontszám számítás
                for c in cars:
                    year, price = c.get("Év"), c.get("Ár_num")
                    if year and price:
                        y = year.split("/")[-1]
                        median = medians.get(y)
                        if median:
                            c["Pontszám"] = round((median - price) / median * 100)
                            c["Medián"] = median

                # Szűrés - MINDEN autót elmentünk, AKÁR nincs is linkje
                for car in cars:
                    link = car.get("Link")
                    
                    # Ha van link, ellenőrizzük
                    if link:
                        if link in seen_links:
                            continue  # Már láttuk
                        seen_links.add(link)
                    else:
                        # Ha nincs link, generálunk egy egyedi azonosítót a cím és ár alapján
                        unique_id = f"no-link-{car.get('Cím', '')}-{car.get('Ár', '')}".strip()
                        if unique_id in seen_links:
                            continue
                        seen_links.add(unique_id)
                    
                    car["Keresés"] = label
                    all_new_cars.append(car)

            context.close()

            print(f"\n📬 Új autók száma ({dealer_id}): {len(all_new_cars)}")

            save_seen_links(dealer_id, seen_links)
            print("💾 Seen linkek frissítve a JSON fájlban.")

            if not all_new_cars:
                send_email(
                    subject=f"🚗 AutoScout – {dealer_id} – nincs új autó",
                    body="Nem találtunk új autókat a megadott keresések alapján.",
                    to_email=emails
                )
                continue

            all_new_cars.sort(key=lambda x: (parse_date(x.get("Év")), x.get("Pontszám") or -999), reverse=True)
            for i, c in enumerate(all_new_cars, 1):
                c["Sorszám"] = i

            filename = f"{dealer_id}_autoscout.xlsx"
            save_to_excel(all_new_cars, filename)

            email_html = build_email_html(all_new_cars, all_medians, dealer_id)
            send_email(
                subject=f"🚗 {len(all_new_cars)} új autó – {dealer_id}",
                body=email_html,
                to_email=emails,
                html=True,
                attachment=filename
            )

        browser.close()
    print("\n✅ SCRAPER KÉSZ")

def parse_date(date_str):
    try:
        month, year = date_str.split("/")
        return int(year), int(month)
    except:
        return (0, 0)

if __name__ == "__main__":
    try:
        run_scraper()
    except Exception as e:
        import traceback
        send_email(
            subject="❌ SCRAPER HIBA",
            body=traceback.format_exc(),
            to_email="aronincze@aronsoft.hu"
        )